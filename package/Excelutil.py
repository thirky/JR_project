#
 # @Date: 2025-02-27 12:16:30
 # @Author: thirky
 # @Description: EXCEL工具类
 #

import openpyxl
import os


#获取excel表格里的站点id和名字
def getSiteIdandName(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    site_id = []
    site_name=[]
    site_region=[]
    bat_num=[]
    for row in range(1,sheet.max_row+1):
        cell_value=sheet[f'A{row}'].value
        if cell_value:
            str(cell_value).strip()
    for row in range(2, sheet.max_row + 1):
        site_id.append(sheet[f'A{row}'].value)
        site_name.append(sheet[f'B{row}'].value)
        site_region.append(sheet[f'C{row}'].value)
        bat_num.append(sheet[f'D{row}'].value)
    return site_id,site_name,site_region,bat_num
