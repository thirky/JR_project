#
 # @Date: 2025-02-27 12:16:30
 # @Author: thirky
 # @Description: EXCEL工具类,注意openpyxl行和列都从1开始计数不是从0开始计数
 #

import openpyxl
import os
from typing import Dict, List, Union

class ExcelUtil:
    """Excel工具类,提供读取Excel文件和处理站点表的功能"""
    DEFAULT_PLATFORMPAGEFORM_PATH = r'.\docs\平台页面站点表.xlsx'
    DEFAULT_JFB_PATH = r'.\docs\佛山移动机房表.xlsx'
    DEFAULT_SITEFORM_PATH = r'.\docs\站点表.xlsx'

    @staticmethod
    def _load_sheet(path: str, sheet_name: str) -> openpyxl.worksheet.worksheet.Worksheet:
        """加载Excel sheet并校验路径和sheet名称"""
        if not os.path.exists(path):
            raise FileNotFoundError(f"Excel文件不存在: {path}")
        workbook = openpyxl.load_workbook(path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' 不存在")
        return workbook[sheet_name]
    
    @classmethod
    def get_site_form(cls, sheet_name: str) -> Dict[str, List[Union[str, float]]]:
        """获取机房表数据(用于平台系统设置-站点管理-新增站点；返回:名称,地址,经度,纬度)"""
        sheet = ExcelUtil._load_sheet(cls.DEFAULT_JFB_PATH, sheet_name)
        return {
            "site_name": [sheet[f'A{row}'].value for row in range(2, sheet.max_row + 1)],
            "site_addr": [sheet[f'B{row}'].value for row in range(2, sheet.max_row + 1)],
            "site_lon": [float(sheet[f'C{row}'].value) for row in range(2, sheet.max_row + 1)],
            "site_lat": [float(sheet[f'D{row}'].value) for row in range(2, sheet.max_row + 1)],
        }

    @classmethod
    def getSiteIdandRegion(cls,sheet_name)-> Dict[str, List[Union[str, float]]]:
        """获取站点ID和区域"""
        sheet = ExcelUtil._load_sheet(cls.DEFAULT_SITEFORM_PATH, sheet_name)
        return {
            "site_id": [sheet[f'A{row}'].value for row in range(2, sheet.max_row + 1)],
            "site_name": [sheet[f'B{row}'].value for row in range(2, sheet.max_row + 1)],
            "site_region": [sheet[f'C{row}'].value for row in range(2, sheet.max_row + 1)],
        }

    @classmethod
    def get_PlatformPage_form(cls,sheet_name: str)-> Dict[str, List[Union[str, float]]]:
        """获取平台页面站点表数据（返回结构化字典）"""
        sheet = ExcelUtil._load_sheet(cls.DEFAULT_PLATFORMPAGEFORM_PATH, sheet_name)
        old = [cell.value.strip() for cell in sheet[1] if cell.value]       
        # 从第2行开始提取数据（使用列表推导式）
        data = {
            "site_region": [sheet[f'A{row}'].value for row in range(2, sheet.max_row + 1)],
            "site_name": [sheet[f'B{row}'].value for row in range(2, sheet.max_row + 1)],
            "site_addr": [sheet[f'C{row}'].value for row in range(2, sheet.max_row + 1)],
            "all_id": [sheet[f'D{row}'].value for row in range(2, sheet.max_row + 1)],
            "power_id": [sheet[f'E{row}'].value for row in range(2, sheet.max_row + 1)],
            "air1_name": [sheet[f'F{row}'].value for row in range(2, sheet.max_row + 1)],
            "air2_name": [sheet[f'G{row}'].value for row in range(2, sheet.max_row + 1)],
            "gw_aid": [sheet[f'H{row}'].value for row in range(2, sheet.max_row + 1)],
            "powerA_mfg": [sheet[f'I{row}'].value for row in range(2, sheet.max_row + 1)],
            "powerA_model": [sheet[f'J{row}'].value for row in range(2, sheet.max_row + 1)],
            "acn_AH": [sheet[f'K{row}'].value for row in range(2, sheet.max_row + 1)],
            "aby_AH": [sheet[f'L{row}'].value for row in range(2, sheet.max_row + 1)],
            "old": old  # 保留原始标题行数据
        }
        return data